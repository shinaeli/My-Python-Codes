import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    work_book = xl.load_workbook(filename)

    # sheet = work_book.get_sheet_by_name('Sheet1')
    sheet = work_book['Sheet1']
    # get_cell = sheet.cell(1, 1)

    # print(sheet.max_row) #4

    for row in range(2, sheet.max_row + 1):
        sheet_cell = sheet.cell(row, 3)
        initial_price = sheet_cell.value
        corrected_price = initial_price * 0.9
        new_cell = sheet.cell(row, 4)
        new_cell.value = corrected_price
        print(new_cell.value)


    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    work_book.save(filename)


process_workbook('transactions.xlsx')






